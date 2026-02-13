# backend/app/services/report_excel.py
from __future__ import annotations

from typing import Any, Dict, List, Optional, Tuple
from openpyxl import load_workbook # type: ignore

from app.config import settings


def _to_str(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _normalize_header(h: str) -> str:
    return _to_str(h).replace("\n", " ").replace("\r", " ").strip()


def _safe_sheet(wb, preferred: str, fallback_index: int) -> Any:
    # try preferred name
    if preferred and preferred in wb.sheetnames:
        return wb[preferred]
    # try case-insensitive match
    lower = {name.lower(): name for name in wb.sheetnames}
    if preferred and preferred.lower() in lower:
        return wb[lower[preferred.lower()]]
    # fallback by index
    names = wb.sheetnames
    if 0 <= fallback_index < len(names):
        return wb[names[fallback_index]]
    # last resort: first sheet
    return wb[wb.sheetnames[0]]


def _read_sheet_as_dicts(path: str, sheet_name: str, fallback_index: int) -> List[Dict[str, Any]]:
    wb = load_workbook(path, data_only=True)
    ws = _safe_sheet(wb, sheet_name, fallback_index)

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [_normalize_header(h) for h in (rows[0] or [])]
    headers = [h for h in headers if h]  # remove empty headers

    items: List[Dict[str, Any]] = []
    for r in rows[1:]:
        if r is None:
            continue
        # build dict by index
        d: Dict[str, Any] = {}
        for idx, h in enumerate(headers):
            if idx < len(r):
                d[h] = r[idx]
            else:
                d[h] = None

        # skip fully empty row
        if all(_to_str(v) == "" for v in d.values()):
            continue

        items.append(d)

    return items


def _possible_keys(d: Dict[str, Any]) -> Dict[str, str]:
    # map normalized header -> real header
    return {_normalize_header(k).lower(): k for k in d.keys()}


def _match_user(row: Dict[str, Any], user: Any) -> bool:
    """
    tries to match row to the logged-in user by one of these (if exists in Excel):
      - personnel code
      - agent code
      - number
      - email azki
      - username
    """
    # user fields (based on your User model)
    user_personnel = _to_str(getattr(user, "personnel_code", ""))
    user_number = _to_str(getattr(user, "number", ""))
    user_email = _to_str(getattr(user, "email_azki", ""))
    user_username = _to_str(getattr(user, "username", ""))

    keys = _possible_keys(row)

    candidates: List[Tuple[str, str]] = [
        ("personnel code", user_personnel),
        ("personnel_code", user_personnel),
        ("agent code", user_personnel),  # sometimes agent code == personnel
        ("agent code", user_number),
        ("agent_code", user_number),
        ("number", user_number),
        ("email azki", user_email),
        ("email_azki", user_email),
        ("user", user_username),
        ("username", user_username),
    ]

    for key_norm, user_val in candidates:
        if not user_val:
            continue
        if key_norm in keys:
            excel_key = keys[key_norm]
            excel_val = _to_str(row.get(excel_key))
            if excel_val and excel_val == user_val:
                return True

    # also try loose numeric compare (e.g. 1156 vs "1156")
    numeric_keys = ["personnel code", "agent code", "number", "agent_code", "personnel_code"]
    user_nums = {user_personnel, user_number}
    user_nums = {x for x in user_nums if x}

    for nk in numeric_keys:
        if nk in keys:
            excel_key = keys[nk]
            excel_val = _to_str(row.get(excel_key))
            if excel_val and excel_val in user_nums:
                return True

    return False


def get_qc_rows_for_user(user: Any) -> List[Dict[str, Any]]:
    items = _read_sheet_as_dicts(
        settings.REPORTS_XLSX_PATH,
        settings.REPORTS_QC_SHEET_NAME,
        fallback_index=0,
    )
    return [r for r in items if _match_user(r, user)]


def get_error_rows_for_user(user: Any) -> List[Dict[str, Any]]:
    items = _read_sheet_as_dicts(
        settings.REPORTS_XLSX_PATH,
        settings.REPORTS_ERRORS_SHEET_NAME,
        fallback_index=1,
    )
    return [r for r in items if _match_user(r, user)]
